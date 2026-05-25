"""Excel 导入导出"""
from fastapi import APIRouter, UploadFile, File, Depends
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ehs_incident.models import SessionLocal, IncidentRecord, Option
from ehs_incident.api.auth import require_admin

router = APIRouter()

# ── 列映射 (Excel表头 → 数据库字段) ──
HEADERS = [
    ("事件单号", "incident_no"), ("异常类型", "incident_type"), ("事件等级", "event_level"),
    ("实验室", "lab"), ("版本", "version"), ("子系统", "subsystem"),
    ("设备名称", "device_name"), ("异常描述", "description"), ("状态", "status"),
    ("报告人", "reporter"), ("报告日期", "report_date"), ("复盘日期", "review_date"),
    ("复盘人", "reviewer"),
]

# 可提取选项的下拉字段
OPTION_FIELDS = ["incident_type", "event_level", "lab", "version", "subsystem", "device_name", "status"]


@router.get("/export/excel")
def export_excel(status: str = "", incident_type: str = "", user=Depends(require_admin)):
    """导出异常事件数据为 Excel"""
    db = SessionLocal()
    try:
        q = db.query(IncidentRecord)
        if status: q = q.filter(IncidentRecord.status == status)
        if incident_type: q = q.filter(IncidentRecord.incident_type == incident_type)
        records = q.order_by(IncidentRecord.id.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "异常事件数据"

        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, (cn, _) in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=cn)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row_idx, record in enumerate(records, 2):
            for col_idx, (_, field) in enumerate(HEADERS, 1):
                val = getattr(record, field, "")
                if val is None: val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

        for col in range(1, len(HEADERS) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from datetime import datetime
        from urllib.parse import quote
        filename = f"EHS异常事件数据_{datetime.now().strftime('%Y%m%d')}.xlsx"
        safe_filename = quote(filename)
        return StreamingResponse(output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=utf-8''{safe_filename}"})
    finally:
        db.close()


@router.post("/import/excel")
async def import_excel(file: UploadFile = File(...), extract_options: bool = True, user=Depends(require_admin)):
    """导入 Excel，自动提取下拉选项"""
    contents = await file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active

    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_map = {}
    header_names = {}
    for cn, db in HEADERS:
        header_names[cn] = db

    for idx, cn in enumerate(header_row):
        if cn:
            cn_clean = str(cn).strip().replace('　', '').replace('\xa0', '')
            if cn_clean in header_names:
                col_map[idx] = header_names[cn_clean]
            else:
                cn_norm = cn_clean.replace(' ', '')
                for hcn, db in header_names.items():
                    if hcn.replace(' ', '') == cn_norm:
                        col_map[idx] = db
                        break

    if not col_map:
        return {"ok": False, "error": "未找到匹配的列标题，请使用标准模板"}

    db = SessionLocal()
    try:
        imported = 0
        extracted_values = {fn: set() for fn in OPTION_FIELDS}

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {}
            for col_idx, db_field in col_map.items():
                val = row[col_idx] if col_idx < len(row) else ""
                if val is None: val = ""
                data[db_field] = str(val).strip()

            if not any(data.values()): continue

            if extract_options:
                for fn in OPTION_FIELDS:
                    if fn in data and data[fn]:
                        extracted_values[fn].add(data[fn])

            if data.get("incident_no"):
                existing = db.query(IncidentRecord).filter(IncidentRecord.incident_no == data["incident_no"]).first()
                if existing:
                    for k, v in data.items():
                        if k in IncidentRecord.__table__.columns.keys() and k != "id":
                            setattr(existing, k, v)
                else:
                    db.add(IncidentRecord(**{k: v for k, v in data.items() if k in IncidentRecord.__table__.columns.keys()}))
                imported += 1
            else:
                db.add(IncidentRecord(**{k: v for k, v in data.items() if k in IncidentRecord.__table__.columns.keys()}))
                imported += 1

        db.commit()

        options_added = 0
        if extract_options:
            for fn, values in extracted_values.items():
                existing = set(o.value for o in db.query(Option).filter(Option.field_name == fn).all())
                for v in values:
                    if v and v not in existing:
                        db.add(Option(field_name=fn, value=v))
                        options_added += 1
            db.commit()

        return {"ok": True, "imported": imported, "options_added": options_added}
    except Exception as e:
        db.rollback()
        return {"ok": False, "error": str(e)}
    finally:
        db.close()


@router.get("/export/template")
def download_template(user=Depends(require_admin)):
    """下载空白导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "异常事件导入模板"

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, (cn, _) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=cn)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from urllib.parse import quote
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=utf-8''" + quote("EHS异常事件导入模板.xlsx")})
