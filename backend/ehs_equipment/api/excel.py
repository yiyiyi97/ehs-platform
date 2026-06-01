"""Excel 导入导出 API"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from ehs_equipment.models import SessionLocal, EquipmentRecord
from ehs_hazard.models import SessionLocal as HazardSessionLocal, User as HazardUser
from ehs_equipment.api.auth import require_admin
from openpyxl import Workbook, load_workbook
from io import BytesIO

router = APIRouter()


def _auth_by_token(token: str):
    """通过 URL query token 认证（复用 hazard 统一用户表）"""
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    db = HazardSessionLocal()
    try:
        user = db.query(HazardUser).filter(HazardUser.token == token).first()
        if not user:
            raise HTTPException(status_code=401, detail="登录已过期")
        return user
    finally:
        db.close()

HEADERS = [
    ("equipment_no", "设备编号"),
    ("equipment_name", "设备名称"),
    ("equipment_type", "设备类型"),
    ("category", "类别"),
    ("lab", "实验室/区域"),
    ("floor", "楼层"),
    ("manufacturer", "制造商"),
    ("model", "型号规格"),
    ("serial_no", "出厂编号"),
    ("location", "安装位置"),
    ("responsible_person", "责任人"),
    ("responsible_dept", "责任部门"),
    ("purchase_date", "采购日期"),
    ("commissioning_date", "投用日期"),
    ("last_check_date", "上次校验日期"),
    ("next_check_date", "下次校验日期"),
    ("check_cycle", "校验周期(月)"),
    ("cert_no", "证书编号"),
    ("status", "设备状态"),
    ("remark", "备注"),
]


@router.get("/export/template")
def download_template(token: str = Query("")):
    _auth_by_token(token)
    wb = Workbook()
    ws = wb.active
    ws.title = "设备台账模板"
    ws.append([h[1] for h in HEADERS])
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=equipment_template.xlsx"})


@router.get("/export/excel")
def export_excel(token: str = Query("")):
    _auth_by_token(token)
    db = SessionLocal()
    try:
        items = db.query(EquipmentRecord).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "设备台账"
        ws.append([h[1] for h in HEADERS])
        for item in items:
            ws.append([getattr(item, h[0], "") for h in HEADERS])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=equipments.xlsx"})
    finally:
        db.close()


@router.post("/import/excel")
def import_excel(data: dict, admin=Depends(require_admin)):
    """前端读取文件后，将每行数据以 JSON 数组形式传入"""
    rows = data.get("rows", [])
    if not rows:
        raise HTTPException(400, "没有数据")

    from ehs_equipment.api.equipments import _calc_next_check_date

    db = SessionLocal()
    try:
        imported = 0
        for row in rows:
            if not row.get("equipment_no") or not row.get("equipment_name"):
                continue
            # 检查是否存在
            existing = db.query(EquipmentRecord).filter(EquipmentRecord.equipment_no == row["equipment_no"]).first()
            if existing:
                continue
            # 自动计算下次校验日期
            if row.get("last_check_date") and row.get("check_cycle"):
                try:
                    row["next_check_date"] = _calc_next_check_date(row["last_check_date"], int(row["check_cycle"]))
                except Exception:
                    pass
            item = EquipmentRecord(**{k: row.get(k, "") for k, _ in HEADERS if k in EquipmentRecord.__table__.columns.keys()})
            db.add(item)
            imported += 1
        db.commit()
        return {"imported": imported}
    except Exception as e:
        db.rollback()
        raise HTTPException(400, str(e))
    finally:
        db.close()
