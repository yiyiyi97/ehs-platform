"""Excel 导入导出"""
from fastapi import APIRouter, UploadFile, File, Depends
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ehs_hazard.models import SessionLocal, HazardRecord, FieldOption
from ehs_hazard.api.auth import require_admin

router = APIRouter()

# ── 列映射 (Excel表头 → 数据库字段) ──
HEADERS = [
    ("隐患单号", "hazard_no"), ("实验室", "lab"), ("版本", "version"), ("楼层", "floor"),
    ("子系统", "subsystem"), ("风险等级", "risk_level"), ("风险类型", "risk_type"),
    ("检查类型", "check_type"), ("相关设备", "equipment"), ("设备阶段", "equipment_phase"),
    ("风险描述", "risk_desc"), ("状态", "status"), ("提出人", "reporter"),
    ("提出日期", "report_date"), ("整改责任人", "responsible_person"),
    ("责任人工号", "responsible_id"), ("责任人部门", "responsible_dept"),
    ("期望闭环日期", "expected_close_date"), ("闭环时间", "close_date"),
    ("验收人员", "verifier"), ("整改建议", "fix_suggestion"),
    ("整改进展", "fix_progress"), ("整改前照片", "photo_before"),
    ("整改后照片", "photo_after"), ("集成接口人", "interface_person"),
    ("是否需要提问题单", "need_issue"), ("具体问题单号", "issue_no"),
    ("是否须在SDE闭环", "need_sde"), ("整改逾期对项目的影响", "overdue_impact"),
]

# 可提取选项的下拉字段
OPTION_FIELDS = ["lab", "version", "floor", "subsystem", "risk_level", "risk_type", "check_type",
                  "equipment", "equipment_phase", "status", "need_issue", "need_sde"]


@router.get("/export/excel")
def export_excel(status: str = "", risk_level: str = "", user=Depends(require_admin)):
    """导出隐患数据为 Excel"""
    db = SessionLocal()
    try:
        q = db.query(HazardRecord)
        if status: q = q.filter(HazardRecord.status == status)
        if risk_level: q = q.filter(HazardRecord.risk_level == risk_level)
        records = q.order_by(HazardRecord.id.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "隐患数据"

        # 表头样式
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 写表头
        for col, (cn, _) in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=cn)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 写数据
        for row_idx, record in enumerate(records, 2):
            for col_idx, (_, field) in enumerate(HEADERS, 1):
                val = getattr(record, field, "")
                if val is None: val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

        # 调整列宽
        for col in range(1, len(HEADERS) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"EHS隐患数据_{__import__('datetime').datetime.now().strftime('%Y%m%d')}.xlsx"
        from urllib.parse import quote
        safe_filename = quote(f"EHS隐患数据_{__import__('datetime').datetime.now().strftime('%Y%m%d')}.xlsx")
        return StreamingResponse(output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=utf-8''{safe_filename}"})
    finally:
        db.close()


@router.post("/import/excel")
async def import_excel(file: UploadFile = File(...), extract_options: bool = True, user=Depends(require_admin)):
    """导入 Excel，自动提取下拉选项"""
    contents = await file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active

    # 读表头，建立 Excel列号 → 数据库字段 映射
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_map = {}  # col_index → db_field
    header_names = {}  # cn_name → db_field
    for cn, db in HEADERS:
        header_names[cn] = db

    for idx, cn in enumerate(header_row):
        if cn:
            cn_clean = str(cn).strip().replace('\u3000', '').replace('\xa0', '')
            if cn_clean in header_names:
                col_map[idx] = header_names[cn_clean]
            else:
                # 模糊匹配：去除空格后比较
                cn_norm = cn_clean.replace(' ', '')
                for hcn, db in header_names.items():
                    if hcn.replace(' ', '') == cn_norm:
                        col_map[idx] = db
                        break

    if not col_map:
        return {"ok": False, "error": "未找到匹配的列标题，请使用标准模板"}

    db = SessionLocal()
    try:
        imported = 0
        extracted_values = {fn: set() for fn in OPTION_FIELDS}  # 收集选项值

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {}
            for col_idx, db_field in col_map.items():
                val = row[col_idx] if col_idx < len(row) else ""
                if val is None: val = ""
                data[db_field] = str(val).strip()

            if not any(data.values()): continue  # 空行跳过

            # 收集选项值
            if extract_options:
                for fn in OPTION_FIELDS:
                    if fn in data and data[fn]:
                        extracted_values[fn].add(data[fn])

            # 更新或创建
            if data.get("hazard_no"):
                existing = db.query(HazardRecord).filter(HazardRecord.hazard_no == data["hazard_no"]).first()
                if existing:
                    for k, v in data.items():
                        if k in HazardRecord.__table__.columns.keys() and k != "id":
                            setattr(existing, k, v)
                else:
                    db.add(HazardRecord(**{k: v for k, v in data.items() if k in HazardRecord.__table__.columns.keys()}))
                imported += 1
            else:
                db.add(HazardRecord(**{k: v for k, v in data.items() if k in HazardRecord.__table__.columns.keys()}))
                imported += 1

        db.commit()

        # 自动提取选项
        options_added = 0
        if extract_options:
            for fn, values in extracted_values.items():
                existing = set(o.value for o in db.query(FieldOption).filter(FieldOption.field_name == fn).all())
                for v in values:
                    if v and v not in existing:
                        db.add(FieldOption(field_name=fn, value=v))
                        options_added += 1
            db.commit()

        return {"ok": True, "imported": imported, "options_added": options_added}
    except Exception as e:
        db.rollback()
        return {"ok": False, "error": str(e)}
    finally:
        db.close()


@router.get("/export/template")
def download_template(user=Depends(require_admin)):
    """下载空白导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "隐患导入模板"

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, (cn, _) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=cn)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from urllib.parse import quote
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=utf-8''" + quote("EHS隐患导入模板.xlsx")})
